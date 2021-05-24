import discord
import random
import asyncio
from discord.utils import get
from discord.ext import commands
import openpyxl


class Coin(commands.Cog, name="코인(Coin)"):

    def __init__(self, app):
        self.app = app

    @commands.command(name='코인등록', help='자신의 아이디를 코인 시스템에 등록합니다.',
                      usage='%코인등록', pass_context=True)
    async def register_coin(self, ctx):
        id = str(ctx.author.id)
        openxl = openpyxl.load_workbook("coin.xlsx")
        wb = openxl.active
        for i in range(1, 100):
            if wb["B" + str(i)].value == id:
                await ctx.channel.send("이미 등록된 아이디입니다.")
                break
            else:
                if wb["B" + str(i)].value == "_":
                    wb["B" + str(i)].value = id
                    wb["A" + str(i)].value = ctx.author.name
                    wb["C" + str(i)].value = 100
                    await ctx.channel.send(str(ctx.author.name) + " 님의 아이디를 등록했습니다.\n"
                                                                  "시작 코인 :coin: 100 개를 지급했습니다.")
                    break
        openxl.save("coin.xlsx")

    @commands.has_permissions(administrator=True)
    @commands.command(name='강제등록', help='타인의 아이디를 코인 시스템에 강제 등록합니다.\n(관리자 권한)',
                      usage='%강제등록', pass_context=True)
    async def register_coin_force(self, ctx, member: discord.Member):
        id = str(member.id)
        openxl = openpyxl.load_workbook("coin.xlsx")
        wb = openxl.active
        for i in range(1, 100):
            if wb["B" + str(i)].value == id:
                await ctx.channel.send("이미 등록된 아이디입니다.")
                break
            else:
                if wb["B" + str(i)].value == "_":
                    wb["B" + str(i)].value = id
                    wb["A" + str(i)].value = member.name
                    await ctx.channel.send(str(member.name) + " 님의 아이디를 등록했습니다.")
                    break
        openxl.save("coin.xlsx")

    @commands.has_permissions(administrator=True)
    @commands.command(name='등록해지', help='타인의 아이디를 코인 시스템에서 삭제합니다.\n(관리자 권한)',
                      usage='%등록해지', pass_context=True)
    async def remove_coin_force(self, ctx, member: discord.Member):
        id = str(member.id)
        openxl = openpyxl.load_workbook("coin.xlsx")
        wb = openxl.active
        for i in range(1, 100):
            if wb["B" + str(i)].value == id:
                wb["B" + str(i)].value = "_"
                await ctx.channel.send(str(member.name) + " 님의 아이디를 해지했습니다.")
                break
        openxl.save("coin.xlsx")

    @commands.command(name='코인명단', help='코인 시스템의 등록자 명단을 확인합니다.',
                      usage='%코인명단', pass_context=True)
    async def register_list(self, ctx):
        openxl = openpyxl.load_workbook("coin.xlsx")
        wb = openxl.active
        embed = discord.Embed(title="<코인 등록자 명단>",
                              description="'%코인등록'을 통해 등록해주세요.")
        for i in range(1, 100):
            if wb["B" + str(i)].value != "_":
                embed.add_field(name=str(wb["A" + str(i)].value), value=":coin:" + str(wb["C" + str(i)].value), inline=False)
        await ctx.send(embed=embed)
        openxl.save("coin.xlsx")

    @commands.has_permissions(administrator=True)
    @commands.command(name='코인설정', help='대상의 코인을 설정합니다.'
                                        '\n(관리자 권한)', usage='%코인설정 @ ~', pass_context=True)
    async def set_coin(self, ctx, member: discord.Member, num):
        id = str(member.id)
        coin = int(num)
        openxl = openpyxl.load_workbook("coin.xlsx")
        wb = openxl.active
        for i in range(1, 100):
            if wb["B" + str(i)].value == id:
                wb["C" + str(i)].value = int(coin)

                await ctx.channel.send(f"코인 설정: {coin}")
                break
        openxl.save("coin.xlsx")

    @commands.has_permissions(administrator=True)
    @commands.command(name='코인증가', help='대상의 코인을 증가시킵니다.'
                                        '\n(관리자 권한)', usage='%코인증가 @ ~', pass_context=True)
    async def gain_coin(self, ctx, member: discord.Member, num):
        id = str(member.id)
        coin = int(num)
        openxl = openpyxl.load_workbook("coin.xlsx")
        wb = openxl.active
        for i in range(1, 100):
            if wb["B" + str(i)].value == id:
                wb["C" + str(i)].value = wb["C" + str(i)].value + int(coin)

                await ctx.channel.send(f"코인 추가: + :coin: {coin}")
                break
        openxl.save("coin.xlsx")

    @commands.has_permissions(administrator=True)
    @commands.command(name='코인감소', help='대상의 코인을 감소시킵니다.'
                                        '\n(관리자 권한)', usage='%코인감소 @ ~', pass_context=True)
    async def lose_coin(self, ctx, member: discord.Member, num):
        id = str(member.id)
        coin = int(num)
        openxl = openpyxl.load_workbook("coin.xlsx")
        wb = openxl.active
        for i in range(1, 100):
            if wb["B" + str(i)].value == id:
                wb["C" + str(i)].value = wb["C" + str(i)].value - int(coin)

                await ctx.channel.send(f"코인 감소: - :coin: {coin}")
                break
        openxl.save("coin.xlsx")

    @commands.command(name='코인', help='자신의 코인을 확인합니다.', usage='%코인')
    async def check_coin(self, ctx):
        id = str(ctx.author.id)
        openxl = openpyxl.load_workbook("coin.xlsx")
        wb = openxl.active
        for i in range(1, 100):
            if wb["B" + str(i)].value == id:
                coin = wb["C" + str(i)].value
                await ctx.channel.send(str(ctx.author.name) + f" 님의 코인: :coin: {coin}")
                break
        openxl.save("coin.xlsx")

    @commands.command(name='베팅', help='자신의 코인을 베팅합니다.'
                                      '\n1/3 확률로 잃거나, 유지하거나, 2배로 반환됩니다.', usage='%코인 ~')
    async def bet_coin(self, ctx, num):
        my_channel = ctx.guild.get_channel(814888257698398289)
        if ctx.channel == my_channel:
            id = str(ctx.author.id)
            openxl = openpyxl.load_workbook("coin.xlsx")
            wb = openxl.active
            for i in range(1, 100):
                if wb["B" + str(i)].value == id:
                    if 0 < int(num) <= wb["C" + str(i)].value:
                        rand = random.randint(-1, 1)
                        wb["C" + str(i)].value = wb["C" + str(i)].value + int(num) * rand
                        embed = discord.Embed(title="<베팅 결과>",
                                              description=ctx.author.name + " 님의 결과")
                        if rand == -1:
                            embed.add_field(name="> 베팅결과", value="-", inline=True)
                        elif rand == 0:
                            embed.add_field(name="> 베팅결과", value="=", inline=True)
                        elif rand == 1:
                            embed.add_field(name="> 베팅결과", value="+", inline=True)
                        embed.add_field(name="> 현재 코인", value=str(wb["C" + str(i)].value), inline=True)
                        await ctx.send(embed=embed)
                    else:
                        await ctx.send("코인이 부족합니다.")
                    break
            openxl.save("coin.xlsx")
        else:
            await ctx.send(":no_entry: 이 채널에서는 사용할 수 없는 명령어입니다.")

    @commands.command(name='트레이드', help='자신과 상대방의 코인을 베팅합니다.'
                                      '\n1/2 확률로 이긴 쪽이 코인을 빼앗아옵니다.', usage='%트레이드 @ ~', pass_context=True)
    async def trade_coin(self, ctx, member: discord.Member, num):
        my_channel = ctx.guild.get_channel(814888257698398289)
        if ctx.channel == my_channel:
            id = str(ctx.author.id)
            oppo_id = str(member.id)
            s_coin = 0
            o_coin = 0
            openxl = openpyxl.load_workbook("coin.xlsx")
            wb = openxl.active
            for i in range(1, 100):
                if wb["B" + str(i)].value == id:
                    s_coin += wb["C" + str(i)].value
                elif wb["B" + str(i)].value == oppo_id:
                    o_coin += wb["C" + str(i)].value
            if s_coin >= int(num) and o_coin >= int(num):
                msg = await ctx.send(ctx.author.name + " 님이 " + member.name + " 님에게 "
                                     + str(num) + " 코인을 걸고 트레이드를 신청합니다."
                                                  "\n수락은 :white_check_mark: 을 눌러주세요.")
                reaction_list = ['✅', '❎']
                for r in reaction_list:
                    await msg.add_reaction(r)

                def check(reaction, user):
                    return str(reaction) in reaction_list and \
                           reaction.message.id == msg.id and user == member

                try:
                    reaction, user = await self.app.wait_for("reaction_add", check=check, timeout=10.0)
                except asyncio.TimeoutError:
                    await msg.edit(content="시간 초과!", delete_after=2)
                else:
                    if str(reaction) == '✅':
                        rand = random.randint(-1, 1)
                        for i in range(1, 100):
                            if wb["B" + str(i)].value == id:
                                if 0 < int(num) <= wb["C" + str(i)].value:
                                    wb["C" + str(i)].value = wb["C" + str(i)].value + int(num) * rand
                                    embed = discord.Embed(title="<트레이드 결과>",
                                                          description=ctx.author.name + " 님의 결과")
                                    if rand == 1:
                                        embed.add_field(name="> 결과", value="승리", inline=True)
                                    elif rand == 0:
                                        embed.add_field(name="> 결과", value="무승부", inline=True)
                                    else:
                                        embed.add_field(name="> 결과", value="패배", inline=True)
                                    embed.add_field(name="> 현재 코인", value=str(wb["C" + str(i)].value), inline=True)
                                    await ctx.send(embed=embed)
                                else:
                                    await ctx.send("코인이 부족합니다.")
                                break
                        for i in range(1, 100):
                            if wb["B" + str(i)].value == oppo_id:
                                if 0 < int(num) <= wb["C" + str(i)].value:
                                    wb["C" + str(i)].value = wb["C" + str(i)].value - int(num) * rand
                                    embed = discord.Embed(title="<트레이드 결과>",
                                                          description=member.name + " 님의 결과")
                                    if rand == -1:
                                        embed.add_field(name="> 결과", value="승리", inline=True)
                                    elif rand == 0:
                                        embed.add_field(name="> 결과", value="무승부", inline=True)
                                    else:
                                        embed.add_field(name="> 결과", value="패배", inline=True)
                                    embed.add_field(name="> 현재 코인", value=str(wb["C" + str(i)].value), inline=True)
                                    await ctx.send(embed=embed)
                                else:
                                    await ctx.send("코인이 부족합니다.")
                                break
                    else:
                        await ctx.send("트레이드 신청을 거절했습니다.")
            else:
                await ctx.send("자신과 상대의 보유 코인 이하로만 베팅할 수 있습니다.")

            openxl.save("coin.xlsx")
        else:
            await ctx.send(":no_entry: 이 채널에서는 사용할 수 없는 명령어입니다.")

    @commands.command(name="코인상점", help='코인 상점의 상품 목록을 공개합니다.', usage='%코인상점')
    async def role_price_list(self, ctx):
        embed = discord.Embed(title="<코인 상점>",
                              description="%구매 (역할 번호) 로 구매해주세요\n"
                                          "'매니저' 역할 보유 시, 50% 할인!")
        openxl_ = openpyxl.load_workbook("Roles.xlsx")
        wb_ = openxl_.active
        for role in ctx.guild.roles:
            if 2 < role.position <= 15:
                embed.add_field(name="> " + role.name, value=":coin: " + str(wb_["B" + str(role.position-2)].value), inline=True)
        await ctx.send(embed=embed)
        openxl_.save("Roles.xlsx")

    @commands.command(name='구매', help='코인을 소비하여 역할을 구매합니다.\n'
                                      '명령어 뒤에 역할 번호를 입력해주세요.', usage='%구매 ~', pass_context=True)
    async def pay_coin(self, ctx, num):
        arole = None
        if 1 <= int(num) <= 13:
            for role in ctx.guild.roles:
                if role.position == int(num) + 2:
                    arole = role
            if arole in ctx.author.roles:
                await ctx.send("이미 " + arole.name + "을(를) 보유하고 있습니다.")
            else:
                openxl_ = openpyxl.load_workbook("Roles.xlsx")
                wb_ = openxl_.active
                role_price = int(wb_["B" + str(num)].value)
                if get(ctx.guild.roles, name='매니저') in ctx.author.roles:
                    role_price = role_price / 2
                id = str(ctx.author.id)
                openxl = openpyxl.load_workbook("coin.xlsx")
                wb = openxl.active
                for i in range(1, 100):
                    if wb["B" + str(i)].value == id:
                        if wb["C" + str(i)].value >= role_price:
                            coin = wb["C" + str(i)].value
                            wb["C" + str(i)].value = coin - role_price
                            await ctx.author.add_roles(arole)
                            await ctx.channel.send(arole.name + "을(를) 구매했습니다!")
                            break
                        else:
                            await ctx.channel.send("코인이 부족합니다.")
                openxl.save("coin.xlsx")
                openxl_.save("Roles.xlsx")
        else:
            await ctx.send("잘못된 값입니다. 1~13 사이의 정수를 입력해주세요.")

    @commands.command(name='판매', help='역할을 판매하여 코인을 획득합니다. (판매가의 10%)\n'
                                      '명령어 뒤에 역할 번호를 입력해주세요.', usage='%판매 ~', pass_context=True)
    async def sell_coin(self, ctx, num):
        arole = None
        if 1 <= int(num) <= 13:
            for role in ctx.guild.roles:
                if role.position == int(num) + 2:
                    arole = role
            else:
                openxl_ = openpyxl.load_workbook("Roles.xlsx")
                wb_ = openxl_.active
                role_price = int(wb_["B" + str(num)].value) // 10
                id = str(ctx.author.id)
                openxl = openpyxl.load_workbook("coin.xlsx")
                wb = openxl.active
                for i in range(1, 100):
                    if wb["B" + str(i)].value == id:
                        if arole in ctx.author.roles:
                            coin = wb["C" + str(i)].value
                            wb["C" + str(i)].value = coin + role_price
                            await ctx.author.remove_roles(arole)
                            await ctx.channel.send(arole.name + "을(를) 판매했습니다. + :coin: " + str(role_price))
                            break
                        else:
                            await ctx.send(arole.name + "을(를) 보유하고 있지 않습니다.")
                openxl.save("coin.xlsx")
                openxl_.save("Roles.xlsx")
        else:
            await ctx.send("잘못된 값입니다. 1~13 사이의 정수를 입력해주세요.")


def setup(app):
    app.add_cog(Coin(app))