import discord
import asyncio
import random
from discord.ext import commands
from discord.utils import get
import openpyxl

class Accounts(commands.Cog, name="계정(Accounts)"):

    def __init__(self, app):
        self.app = app

    @commands.command(name='회원가입', help='0군 계정을 생성합니다.',
                      usage='%회원가입', pass_context=True)
    async def sign_in(self, ctx):
        id = str(ctx.author.id)
        openxl = openpyxl.load_workbook("account.xlsx")
        wb = openxl.active
        for i in range(1, 100):
            if wb["B" + str(i)].value == id:
                await ctx.channel.send("이미 등록된 아이디입니다.")
            elif wb["B" + str(i)].value == "_":
                dm_channel = await ctx.author.create_dm()
                await dm_channel.send("비밀번호를 입력해주세요.")
                def check(m):
                    return m.author == ctx.author and m.channel == dm_channel
                try:
                    msg = await self.app.wait_for("message", check=check, timeout=10.0)
                except asyncio.TimeoutError:
                    await dm_channel.send("시간 초과되었습니다.")
                    break
                else:
                    wb["C" + str(i)].value = str(msg)
                    wb["B" + str(i)].value = id
                    wb["A" + str(i)].value = ctx.author.name
                    await dm_channel.send(str(ctx.author.name) + " 님의 아이디를 등록했습니다. ")
                    break
        openxl.save("account.xlsx")

    @commands.command(name='비밀번호변경', help='0군 계정의 비밀번호를 재설정합니다.',
                      usage='%비밀번호변경 ~', pass_context=True)
    async def password_change(self, ctx, password):
        if get(ctx.guild.roles, name='로그인') in ctx.author.roles:
            id = str(ctx.author.id)
            openxl = openpyxl.load_workbook("account.xlsx")
            wb = openxl.active
            dm_channel = await ctx.author.create_dm()
            for i in range(1, 100):
                if wb["B" + str(i)].value == id:
                    wb["C" + str(i)].value = password
                    await dm_channel.send(":white_check_mark: 변경되었습니다.")
            openxl.save("account.xlsx")
        else:
            await ctx.send("로그인 되어 있지 않습니다.")

    @commands.command(name='로그인', help='0군 계정으로 로그인합니다.',
                      usage='%로그인', pass_context=True)
    async def log_in(self, ctx):
        dm_channel = await ctx.author.create_dm()
        if get(ctx.guild.roles, name='로그인') in ctx.author.roles:
            await dm_channel.send("이미 로그인되어 있습니다.")
        else:
            id = str(ctx.author.id)
            openxl = openpyxl.load_workbook("account.xlsx")
            wb = openxl.active

            for i in range(1, 100):
                if wb["B" + str(i)].value == id:
                    await dm_channel.send("비밀번호를 입력해주세요.")

                    def check(m):
                        return m.author == ctx.author and m.channel == dm_channel

                    try:
                        msg = await self.app.wait_for("message", check=check, timeout=10.0)
                    except asyncio.TimeoutError:
                        await dm_channel.send("시간 초과되었습니다.")
                    else:
                        if wb["C" + str(i)].value == msg:
                            await dm_channel.send(str(ctx.author) + " 님이 로그인했습니다.",
                                                  reason=dm_channel.send(str(ctx.author) + " 님이 로그인했습니다."))
                            break
            openxl.save("account.xlsx")

    @commands.command(name='코인등록', help='자신의 아이디를 코인 시스템에 등록합니다.',
                      usage='%코인등록', pass_context=True)
    async def register_coin(self, ctx):
        if get(ctx.guild.roles, name='로그인') in ctx.author.roles:
            id = str(ctx.author.id)
            openxl = openpyxl.load_workbook("coin.xlsx")
            wb = openxl.active
            for i in range(1, 100):
                if wb["B" + str(i)].value == id:
                    await ctx.channel.send("이미 등록된 아이디입니다.")
                    break
                else:
                    if wb["B" + str(i)].value == "_":
                        wb["B" + str(i)].value = id
                        wb["A" + str(i)].value = ctx.author.name
                        wb["C" + str(i)].value = 0
                        await ctx.channel.send(str(ctx.author.name) + " 님의 아이디를 등록했습니다. ")
                        break
            openxl.save("coin.xlsx")
        else:
            await ctx.send("로그인 되어 있지 않습니다.")

def setup(app):
    app.add_cog(Accounts(app))